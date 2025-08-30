"""
File generation utilities for PDF, Excel, and CSV documents.

This module provides enhanced generators that build upon the existing
implementations in the ISP Framework while adding additional features like
template integration, better styling, and multi-tenant support.
"""

import csv
import io
import logging
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, BinaryIO, Dict, List, Optional, Tuple, Union

# Excel imports
import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart
from openpyxl.styles import Alignment, Border, Fill, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing

# ReportLab imports for PDF generation
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, legal, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, inch
from reportlab.platypus import (
    HRFlowable,
    Image,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.platypus.flowables import Flowable

logger = logging.getLogger(__name__)


@dataclass
class DocumentMetadata:
    """Metadata for generated documents."""

    file_id: str
    original_filename: str
    file_type: str
    file_size: int
    generated_at: datetime
    tenant_id: str
    template_name: Optional[str] = None
    template_version: Optional[str] = None
    generator_version: str = "1.0.0"
    custom_metadata: Dict[str, Any] = None

    def __post_init__(self):
        if self.custom_metadata is None:
            self.custom_metadata = {}


class PDFGenerator:
    """Enhanced PDF generation with template support and advanced styling."""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize PDF generator with configuration."""
        self.config = config or {}
        self.default_page_size = getattr(
            __import__(
                "reportlab.lib.pagesizes",
                fromlist=[self.config.get("page_size", "letter")],
            ),
            self.config.get("page_size", "letter"),
            letter,
        )
        self.default_margins = self.config.get("margins", 72)
        self.font_family = self.config.get("font_family", "Helvetica")
        self.font_size = self.config.get("font_size", 12)

        # Initialize styles
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom paragraph styles."""
        # Company header style
        self.styles.add(
            ParagraphStyle(
                name="CompanyHeader",
                parent=self.styles["Heading1"],
                fontSize=24,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue,
                fontName="Helvetica-Bold",
            )
        )

        # Section header style
        self.styles.add(
            ParagraphStyle(
                name="SectionHeader",
                parent=self.styles["Heading2"],
                fontSize=14,
                spaceAfter=12,
                spaceBefore=20,
                alignment=TA_LEFT,
                textColor=colors.darkblue,
                fontName="Helvetica-Bold",
                borderWidth=1,
                borderColor=colors.darkblue,
                borderPadding=5,
                backColor=colors.lightgrey,
            )
        )

        # Invoice total style
        self.styles.add(
            ParagraphStyle(
                name="InvoiceTotal",
                parent=self.styles["Normal"],
                fontSize=14,
                fontName="Helvetica-Bold",
                alignment=TA_RIGHT,
                textColor=colors.darkblue,
                spaceAfter=10,
            )
        )

        # Footer style
        self.styles.add(
            ParagraphStyle(
                name="Footer",
                parent=self.styles["Normal"],
                fontSize=9,
                alignment=TA_CENTER,
                textColor=colors.grey,
                spaceAfter=0,
            )
        )

    def generate_invoice(
        self,
        invoice_data: Dict[str, Any],
        output_path: Optional[str] = None,
        template_name: str = "standard_invoice",
        tenant_id: Optional[str] = None,
    ) -> Tuple[str, DocumentMetadata]:
        """
        Generate invoice PDF with enhanced styling and branding.

        Args:
            invoice_data: Invoice data dictionary
            output_path: Output file path (optional)
            template_name: Template to use for invoice
            tenant_id: Tenant ID for multi-tenant support

        Returns:
            Tuple of (file_path, metadata)
        """
        if not output_path:
            file_id = str(uuid.uuid4())
            output_path = f"/tmp/invoice_{invoice_data.get('invoice_number', file_id)}_{file_id}.pdf"
        else:
            file_id = str(uuid.uuid4())

        # Create PDF document
        doc = SimpleDocTemplate(
            output_path,
            pagesize=self.default_page_size,
            rightMargin=self.default_margins,
            leftMargin=self.default_margins,
            topMargin=self.default_margins,
            bottomMargin=self.default_margins,
        )

        # Build content
        story = []

        # Add company header
        self._add_company_header(story, invoice_data.get("company_info", {}))

        # Add invoice header
        self._add_invoice_header(story, invoice_data)

        # Add customer information
        self._add_customer_info(story, invoice_data.get("customer_info", {}))

        # Add invoice items
        self._add_invoice_items(story, invoice_data)

        # Add payment information
        self._add_payment_info(story, invoice_data.get("payment_info", {}))

        # Add footer
        self._add_footer(story, invoice_data)

        # Build PDF
        doc.build(story)

        # Get file size
        file_size = Path(output_path).stat().st_size

        # Create metadata
        metadata = DocumentMetadata(
            file_id=file_id,
            original_filename=f"invoice_{invoice_data.get('invoice_number', 'unknown')}.pdf",
            file_type="application/pdf",
            file_size=file_size,
            generated_at=datetime.now(timezone.utc),
            tenant_id=tenant_id or "default",
            template_name=template_name,
            custom_metadata={
                "invoice_number": invoice_data.get("invoice_number"),
                "customer_id": invoice_data.get("customer_id"),
                "total_amount": invoice_data.get("total_amount"),
            },
        )

        logger.info(f"Generated invoice PDF: {output_path} (size: {file_size} bytes)")
        return output_path, metadata

    def generate_report(
        self,
        report_data: Dict[str, Any],
        output_path: Optional[str] = None,
        template_name: str = "standard_report",
        tenant_id: Optional[str] = None,
    ) -> Tuple[str, DocumentMetadata]:
        """
        Generate report PDF with charts and tables.

        Args:
            report_data: Report data dictionary
            output_path: Output file path (optional)
            template_name: Template to use for report
            tenant_id: Tenant ID for multi-tenant support

        Returns:
            Tuple of (file_path, metadata)
        """
        if not output_path:
            file_id = str(uuid.uuid4())
            report_name = report_data.get("title", "report").replace(" ", "_").lower()
            output_path = f"/tmp/{report_name}_{file_id}.pdf"
        else:
            file_id = str(uuid.uuid4())

        doc = SimpleDocTemplate(
            output_path,
            pagesize=self.default_page_size,
            rightMargin=self.default_margins,
            leftMargin=self.default_margins,
            topMargin=self.default_margins,
            bottomMargin=self.default_margins,
        )

        story = []

        # Add report header
        title = report_data.get("title", "Report")
        story.append(Paragraph(title, self.styles["CompanyHeader"]))
        story.append(Spacer(1, 20))

        # Add report metadata
        generated_date = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        story.append(Paragraph(f"Generated: {generated_date}", self.styles["Normal"]))

        if report_data.get("date_range"):
            date_range = report_data["date_range"]
            story.append(
                Paragraph(
                    f"Period: {date_range.get('start')} to {date_range.get('end')}",
                    self.styles["Normal"],
                )
            )

        story.append(Spacer(1, 20))

        # Add executive summary if provided
        if report_data.get("summary"):
            story.append(Paragraph("Executive Summary", self.styles["SectionHeader"]))
            story.append(Paragraph(report_data["summary"], self.styles["Normal"]))
            story.append(Spacer(1, 20))

        # Process report sections
        sections = report_data.get("sections", [])
        for section in sections:
            self._add_report_section(story, section)

        # Add charts if provided
        if report_data.get("charts"):
            for chart_data in report_data["charts"]:
                self._add_chart_to_story(story, chart_data)

        # Build PDF
        doc.build(story)

        # Get file size
        file_size = Path(output_path).stat().st_size

        # Create metadata
        metadata = DocumentMetadata(
            file_id=file_id,
            original_filename=f"{report_data.get('title', 'report').replace(' ', '_').lower()}.pdf",
            file_type="application/pdf",
            file_size=file_size,
            generated_at=datetime.now(timezone.utc),
            tenant_id=tenant_id or "default",
            template_name=template_name,
            custom_metadata={
                "title": report_data.get("title"),
                "sections_count": len(sections),
                "charts_count": len(report_data.get("charts", [])),
            },
        )

        logger.info(f"Generated report PDF: {output_path} (size: {file_size} bytes)")
        return output_path, metadata

    def _add_company_header(self, story: List, company_info: Dict[str, Any]):
        """Add company header to document."""
        company_name = company_info.get("name", "DotMac ISP")
        story.append(Paragraph(company_name, self.styles["CompanyHeader"]))

        if company_info.get("logo_path"):
            # Add logo if provided
            try:
                logo = Image(company_info["logo_path"], width=2 * inch, height=1 * inch)
                story.append(logo)
                story.append(Spacer(1, 12))
            except Exception as e:
                logger.warning(f"Could not add company logo: {e}")

        # Company details
        details = []
        for field in ["address", "city", "phone", "email", "website"]:
            if company_info.get(field):
                details.append(company_info[field])

        if details:
            company_details = "<br/>".join(details)
            story.append(Paragraph(company_details, self.styles["Normal"]))
            story.append(Spacer(1, 20))

    def _add_invoice_header(self, story: List, invoice_data: Dict[str, Any]):
        """Add invoice header information."""
        story.append(Paragraph("INVOICE", self.styles["CompanyHeader"]))
        story.append(Spacer(1, 20))

        # Invoice info table
        invoice_info = []

        required_fields = [
            ("Invoice Number:", "invoice_number"),
            ("Invoice Date:", "invoice_date"),
            ("Due Date:", "due_date"),
        ]

        optional_fields = [
            ("Customer ID:", "customer_id"),
            ("Purchase Order:", "po_number"),
            ("Terms:", "payment_terms"),
        ]

        # Add required fields
        for label, key in required_fields:
            value = invoice_data.get(key, "N/A")
            invoice_info.append([label, str(value)])

        # Add optional fields if present
        for label, key in optional_fields:
            if invoice_data.get(key):
                invoice_info.append([label, str(invoice_data[key])])

        invoice_table = Table(invoice_info, colWidths=[2 * inch, 3 * inch])
        invoice_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(invoice_table)
        story.append(Spacer(1, 20))

    def _add_customer_info(self, story: List, customer_info: Dict[str, Any]):
        """Add customer billing information."""
        if not customer_info:
            return

        story.append(Paragraph("Bill To:", self.styles["SectionHeader"]))

        # Format customer address
        bill_to_lines = []

        if customer_info.get("name"):
            bill_to_lines.append(customer_info["name"])

        if customer_info.get("company"):
            bill_to_lines.append(customer_info["company"])

        # Address lines
        if customer_info.get("address"):
            bill_to_lines.append(customer_info["address"])

        if customer_info.get("address2"):
            bill_to_lines.append(customer_info["address2"])

        # City, state, zip
        city_line_parts = []
        for field in ["city", "state", "zip_code"]:
            if customer_info.get(field):
                city_line_parts.append(customer_info[field])

        if city_line_parts:
            bill_to_lines.append(" ".join(city_line_parts))

        if customer_info.get("country"):
            bill_to_lines.append(customer_info["country"])

        # Contact info
        if customer_info.get("email"):
            bill_to_lines.append(f"Email: {customer_info['email']}")

        if customer_info.get("phone"):
            bill_to_lines.append(f"Phone: {customer_info['phone']}")

        bill_to_text = "<br/>".join(bill_to_lines)
        story.append(Paragraph(bill_to_text, self.styles["Normal"]))
        story.append(Spacer(1, 20))

    def _add_invoice_items(self, story: List, invoice_data: Dict[str, Any]):
        """Add invoice line items table."""
        story.append(Paragraph("Invoice Items:", self.styles["SectionHeader"]))

        # Prepare table data
        items_data = [["Description", "Quantity", "Unit Price", "Total"]]

        items = invoice_data.get("items", [])
        for item in items:
            items_data.append(
                [
                    item.get("description", ""),
                    str(item.get("quantity", 0)),
                    f"${float(item.get('unit_price', 0)):.2f}",
                    f"${float(item.get('total', 0)):.2f}",
                ]
            )

        # Add subtotals
        subtotal = float(invoice_data.get("subtotal", 0))
        tax_amount = float(invoice_data.get("tax_amount", 0))
        discount_amount = float(invoice_data.get("discount_amount", 0))
        total_amount = float(invoice_data.get("total_amount", 0))

        # Add empty row for separation
        items_data.append(["", "", "", ""])

        # Add financial summary rows
        if subtotal > 0:
            items_data.append(["", "", "Subtotal:", f"${subtotal:.2f}"])

        if discount_amount > 0:
            items_data.append(["", "", "Discount:", f"-${discount_amount:.2f}"])

        if tax_amount > 0:
            items_data.append(["", "", "Tax:", f"${tax_amount:.2f}"])

        items_data.append(["", "", "Total:", f"${total_amount:.2f}"])

        # Create table
        items_table = Table(
            items_data, colWidths=[3.5 * inch, 1 * inch, 1.5 * inch, 1.5 * inch]
        )

        # Calculate number of item rows (excluding summary rows)
        item_rows = len(items)

        items_table.setStyle(
            TableStyle(
                [
                    # Header row styling
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    # Item rows styling
                    ("BACKGROUND", (0, 1), (-1, item_rows), colors.beige),
                    (
                        "ALIGN",
                        (0, 1),
                        (0, item_rows),
                        "LEFT",
                    ),  # Description left-aligned
                    ("GRID", (0, 0), (-1, item_rows), 1, colors.black),
                    # Summary rows styling
                    ("BACKGROUND", (0, item_rows + 2), (-1, -1), colors.lightgrey),
                    ("FONTNAME", (0, item_rows + 2), (-1, -1), "Helvetica-Bold"),
                    ("ALIGN", (2, item_rows + 2), (-1, -1), "RIGHT"),
                    ("GRID", (0, item_rows + 2), (-1, -1), 1, colors.black),
                    # Total row special styling
                    ("BACKGROUND", (0, -1), (-1, -1), colors.darkblue),
                    ("TEXTCOLOR", (0, -1), (-1, -1), colors.whitesmoke),
                    ("FONTSIZE", (0, -1), (-1, -1), 12),
                ]
            )
        )

        story.append(items_table)
        story.append(Spacer(1, 30))

    def _add_payment_info(self, story: List, payment_info: Dict[str, Any]):
        """Add payment information section."""
        if not payment_info:
            return

        story.append(Paragraph("Payment Information:", self.styles["SectionHeader"]))

        payment_lines = []

        if payment_info.get("terms"):
            payment_lines.append(f"Payment Terms: {payment_info['terms']}")

        if payment_info.get("methods"):
            methods = payment_info["methods"]
            if isinstance(methods, list):
                methods_text = ", ".join(methods)
            else:
                methods_text = str(methods)
            payment_lines.append(f"Payment Methods: {methods_text}")

        if payment_info.get("bank_details"):
            payment_lines.append("Bank Details:")
            bank_details = payment_info["bank_details"]
            for key, value in bank_details.items():
                payment_lines.append(f"  {key.replace('_', ' ').title()}: {value}")

        if payment_info.get("notes"):
            payment_lines.append(f"Notes: {payment_info['notes']}")

        if payment_lines:
            payment_text = "<br/>".join(payment_lines)
            story.append(Paragraph(payment_text, self.styles["Normal"]))
            story.append(Spacer(1, 20))

    def _add_footer(self, story: List, invoice_data: Dict[str, Any]):
        """Add footer to document."""
        story.append(HRFlowable(width="100%", thickness=1, color=colors.black))
        story.append(Spacer(1, 10))

        footer_text = invoice_data.get("footer_text", "Thank you for your business!")
        story.append(Paragraph(footer_text, self.styles["Footer"]))

        # Add disclaimer if provided
        if invoice_data.get("disclaimer"):
            story.append(Spacer(1, 5))
            story.append(
                Paragraph(f"<i>{invoice_data['disclaimer']}</i>", self.styles["Footer"])
            )

    def _add_report_section(self, story: List, section: Dict[str, Any]):
        """Add a report section to the story."""
        section_type = section.get("type", "paragraph")

        if section_type == "heading":
            story.append(
                Paragraph(section.get("content", ""), self.styles["SectionHeader"])
            )
            story.append(Spacer(1, 12))

        elif section_type == "paragraph":
            story.append(Paragraph(section.get("content", ""), self.styles["Normal"]))
            story.append(Spacer(1, 12))

        elif section_type == "table":
            self._add_data_table(
                story, section.get("data", []), section.get("headers", [])
            )

        elif section_type == "list":
            items = section.get("items", [])
            for item in items:
                story.append(Paragraph(f"• {item}", self.styles["Normal"]))
            story.append(Spacer(1, 12))

    def _add_data_table(self, story: List, data: List[List], headers: List[str] = None):
        """Add a data table to the story."""
        if not data:
            return

        table_data = []
        if headers:
            table_data.append(headers)
        table_data.extend(data)

        table = Table(table_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0 if headers else -1), colors.grey),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0 if headers else -1),
                        colors.whitesmoke,
                    ),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0 if headers else -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0 if headers else -1), 12),
                    ("BACKGROUND", (0, 1 if headers else 0), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 12))

    def _add_chart_to_story(self, story: List, chart_data: Dict[str, Any]):
        """Add a chart to the story (placeholder for future chart integration)."""
        # For now, add a placeholder
        chart_title = chart_data.get("title", "Chart")
        story.append(Paragraph(f"Chart: {chart_title}", self.styles["SectionHeader"]))
        story.append(
            Paragraph(
                f"Chart data: {len(chart_data.get('data', []))} data points",
                self.styles["Normal"],
            )
        )
        story.append(Spacer(1, 20))


class ExcelGenerator:
    """Enhanced Excel generation with advanced styling and chart support."""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize Excel generator with configuration."""
        self.config = config or {}
        self.default_sheet_name = self.config.get("default_sheet_name", "Data")
        self.auto_fit_columns = self.config.get("auto_fit_columns", True)
        self.freeze_panes = self.config.get("freeze_panes", True)

    def generate_report(
        self,
        data: List[Dict[str, Any]],
        output_path: Optional[str] = None,
        sheet_name: Optional[str] = None,
        template_config: Optional[Dict[str, Any]] = None,
        tenant_id: Optional[str] = None,
    ) -> Tuple[str, DocumentMetadata]:
        """
        Generate Excel report with advanced styling and formulas.

        Args:
            data: List of data dictionaries
            output_path: Output file path (optional)
            sheet_name: Excel sheet name
            template_config: Template configuration for styling
            tenant_id: Tenant ID for multi-tenant support

        Returns:
            Tuple of (file_path, metadata)
        """
        if not output_path:
            file_id = str(uuid.uuid4())
            output_path = f"/tmp/excel_report_{file_id}.xlsx"
        else:
            file_id = str(uuid.uuid4())

        sheet_name = sheet_name or self.default_sheet_name
        template_config = template_config or {}

        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        if not data:
            # Create empty workbook
            workbook.save(output_path)
            file_size = Path(output_path).stat().st_size

            metadata = DocumentMetadata(
                file_id=file_id,
                original_filename=f"empty_report.xlsx",
                file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                file_size=file_size,
                generated_at=datetime.now(timezone.utc),
                tenant_id=tenant_id or "default",
                custom_metadata={"rows": 0, "columns": 0},
            )

            return output_path, metadata

        # Get column names
        columns = list(data[0].keys())

        # Setup styles
        self._setup_excel_styles(workbook, template_config)

        # Write headers
        for col_idx, column in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column.replace("_", " ").title()
            cell.style = "Header"

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, column in enumerate(columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = row_data.get(column)

                # Handle different data types
                if isinstance(value, (datetime)):
                    cell.value = value
                    cell.number_format = "YYYY-MM-DD HH:MM:SS"
                elif isinstance(value, Decimal):
                    cell.value = float(value)
                    cell.number_format = "#,##0.00"
                elif isinstance(value, (int, float)):
                    cell.value = value
                    if column.lower() in ["amount", "price", "cost", "total"]:
                        cell.number_format = "#,##0.00"
                elif isinstance(value, bool):
                    cell.value = "Yes" if value else "No"
                else:
                    cell.value = str(value) if value is not None else ""

                # Apply data cell style
                if row_idx % 2 == 0:
                    cell.style = "DataEven"
                else:
                    cell.style = "DataOdd"

        # Auto-fit columns
        if self.auto_fit_columns:
            self._auto_fit_columns(worksheet)

        # Freeze panes
        if self.freeze_panes:
            worksheet.freeze_panes = "A2"

        # Add filters
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}{len(data) + 1}"
        )

        # Add charts if configured
        if template_config.get("charts"):
            self._add_excel_charts(
                worksheet, template_config["charts"], len(data), len(columns)
            )

        # Save workbook
        workbook.save(output_path)

        # Get file size
        file_size = Path(output_path).stat().st_size

        # Create metadata
        metadata = DocumentMetadata(
            file_id=file_id,
            original_filename=f"report.xlsx",
            file_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_size=file_size,
            generated_at=datetime.now(timezone.utc),
            tenant_id=tenant_id or "default",
            custom_metadata={
                "rows": len(data),
                "columns": len(columns),
                "sheet_name": sheet_name,
            },
        )

        logger.info(
            f"Generated Excel report: {output_path} ({len(data)} rows, {len(columns)} columns)"
        )
        return output_path, metadata

    def _setup_excel_styles(
        self, workbook: openpyxl.Workbook, template_config: Dict[str, Any]
    ):
        """Setup custom styles for the Excel workbook."""
        # Header style
        header_style = NamedStyle(name="Header")
        header_style.font = Font(bold=True, color="FFFFFF", size=12)
        header_style.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(header_style)

        # Data styles
        data_even_style = NamedStyle(name="DataEven")
        data_even_style.font = Font(size=10)
        data_even_style.fill = PatternFill(
            start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
        )
        data_even_style.alignment = Alignment(horizontal="left", vertical="center")
        data_even_style.border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        workbook.add_named_style(data_even_style)

        data_odd_style = NamedStyle(name="DataOdd")
        data_odd_style.font = Font(size=10)
        data_odd_style.fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        data_odd_style.alignment = Alignment(horizontal="left", vertical="center")
        data_odd_style.border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        workbook.add_named_style(data_odd_style)

    def _auto_fit_columns(self, worksheet):
        """Auto-fit column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_excel_charts(
        self,
        worksheet,
        chart_configs: List[Dict[str, Any]],
        data_rows: int,
        data_cols: int,
    ):
        """Add charts to the Excel worksheet."""
        chart_row_offset = data_rows + 5

        for i, chart_config in enumerate(chart_configs):
            chart_type = chart_config.get("type", "bar")
            chart_title = chart_config.get("title", f"Chart {i+1}")

            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                continue  # Skip unknown chart types

            chart.title = chart_title
            chart.height = 10
            chart.width = 15

            # Position chart
            chart_cell = f"A{chart_row_offset + i * 20}"
            worksheet.add_chart(chart, chart_cell)


class CSVGenerator:
    """Enhanced CSV generation with proper encoding and streaming support."""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize CSV generator with configuration."""
        self.config = config or {}
        self.encoding = self.config.get("encoding", "utf-8")
        self.delimiter = self.config.get("delimiter", ",")
        self.quote_char = self.config.get("quote_char", '"')
        self.line_terminator = self.config.get("line_terminator", "\r\n")

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        output_path: Optional[str] = None,
        columns: Optional[List[str]] = None,
        tenant_id: Optional[str] = None,
    ) -> Tuple[str, DocumentMetadata]:
        """
        Export data to CSV file with proper formatting.

        Args:
            data: List of data dictionaries
            output_path: Output file path (optional)
            columns: Column names to include (optional)
            tenant_id: Tenant ID for multi-tenant support

        Returns:
            Tuple of (file_path, metadata)
        """
        if not output_path:
            file_id = str(uuid.uuid4())
            output_path = f"/tmp/export_{file_id}.csv"
        else:
            file_id = str(uuid.uuid4())

        # Determine columns
        if not columns and data:
            columns = list(data[0].keys())
        elif not columns:
            columns = []

        with open(output_path, "w", newline="", encoding=self.encoding) as csvfile:
            writer = csv.DictWriter(
                csvfile,
                fieldnames=columns,
                delimiter=self.delimiter,
                quotechar=self.quote_char,
                lineterminator=self.line_terminator,
            )

            if columns:
                writer.writeheader()

            for row in data:
                # Clean row data
                cleaned_row = {}
                for key in columns:
                    value = row.get(key, "")
                    if isinstance(value, datetime):
                        cleaned_row[key] = value.isoformat()
                    elif isinstance(value, Decimal):
                        cleaned_row[key] = str(value)
                    elif isinstance(value, (dict, list)):
                        cleaned_row[key] = str(value)
                    elif value is None:
                        cleaned_row[key] = ""
                    else:
                        cleaned_row[key] = value

                writer.writerow(cleaned_row)

        # Get file size
        file_size = Path(output_path).stat().st_size

        # Create metadata
        metadata = DocumentMetadata(
            file_id=file_id,
            original_filename="export.csv",
            file_type="text/csv",
            file_size=file_size,
            generated_at=datetime.now(timezone.utc),
            tenant_id=tenant_id or "default",
            custom_metadata={
                "rows": len(data),
                "columns": len(columns) if columns else 0,
                "encoding": self.encoding,
            },
        )

        logger.info(
            f"Generated CSV export: {output_path} ({len(data)} rows, {len(columns)} columns)"
        )
        return output_path, metadata
