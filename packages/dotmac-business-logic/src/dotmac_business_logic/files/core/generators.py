"""
File generators for various formats (PDF, Excel, CSV)
Handles document generation with optional dependencies for graceful degradation
"""

import warnings
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Optional, Union

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    warnings.warn(
        "pandas not available - Excel/CSV generation will be limited", stacklevel=2
    )

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    letter = (612, 792)  # Default letter size in points
    A4 = (595, 842)  # Default A4 size in points
    warnings.warn(
        "reportlab not available - PDF generation will use fallback methods",
        stacklevel=2,
    )

try:
    from openpyxl.styles import Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn(
        "openpyxl not available - Excel generation will use CSV fallback", stacklevel=2
    )


@dataclass
class DocumentMetadata:
    """Document metadata for generated files."""

    title: str
    author: str = "DotMAC System"
    subject: str = ""
    creator: str = "DotMAC File Generator"
    created_date: datetime = None

    def __post_init__(self):
        if self.created_date is None:
            self.created_date = datetime.now()


class BaseGenerator:
    """Base class for all file generators."""

    def __init__(self, metadata: Optional[DocumentMetadata] = None):
        self.metadata = metadata or DocumentMetadata(title="Generated Document")

    def _ensure_path(self, output_path: Optional[Union[str, Path]]) -> Optional[Path]:
        """Ensure output path exists if provided."""
        if output_path:
            path = Path(output_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            return path
        return None


class PDFGenerator(BaseGenerator):
    """PDF document generator with graceful degradation."""

    def __init__(
        self, metadata: Optional[DocumentMetadata] = None, page_size: tuple = letter
    ):
        super().__init__(metadata)
        self.page_size = page_size
        self._check_dependencies()

    def _check_dependencies(self):
        """Check if required dependencies are available."""
        if not REPORTLAB_AVAILABLE:
            warnings.warn(
                "reportlab not available - PDF generation will create text files instead",
                stacklevel=2,
            )

    def generate_simple_pdf(
        self, content: str, output_path: Optional[Union[str, Path]] = None
    ) -> bytes:
        """Generate a simple PDF from text content."""
        if REPORTLAB_AVAILABLE:
            return self._generate_reportlab_pdf(content, output_path)
        else:
            return self._generate_fallback_pdf(content, output_path)

    def _generate_reportlab_pdf(
        self, content: str, output_path: Optional[Union[str, Path]]
    ) -> bytes:
        """Generate PDF using reportlab."""
        buffer = BytesIO()

        # Create document
        if output_path:
            path = self._ensure_path(output_path)
            doc = SimpleDocTemplate(str(path), pagesize=self.page_size)
        else:
            doc = SimpleDocTemplate(buffer, pagesize=self.page_size)

        # Create styles
        styles = getSampleStyleSheet()
        normal_style = styles["Normal"]
        title_style = styles["Title"]

        # Build content
        story = []

        # Add title
        if self.metadata.title:
            story.append(Paragraph(self.metadata.title, title_style))
            story.append(Spacer(1, 12))

        # Add content paragraphs
        for line in content.split("\n"):
            if line.strip():
                story.append(Paragraph(line, normal_style))
            else:
                story.append(Spacer(1, 12))

        # Build PDF
        doc.build(story)

        if output_path:
            with open(output_path, "rb") as f:
                return f.read()
        else:
            return buffer.getvalue()

    def _generate_fallback_pdf(
        self, content: str, output_path: Optional[Union[str, Path]]
    ) -> bytes:
        """Generate text file as fallback when reportlab unavailable."""
        warnings.warn(
            "PDF generation unavailable - creating text file instead", stacklevel=2
        )

        # Create formatted text content
        formatted_content = f"""
{self.metadata.title}
{"=" * len(self.metadata.title)}

Generated: {self.metadata.created_date}
Author: {self.metadata.author}

{content}
"""

        # Save or return content
        if output_path:
            path = self._ensure_path(output_path)
            # Change extension to .txt for fallback
            path = path.with_suffix(".txt")
            with open(path, "w", encoding="utf-8") as f:
                f.write(formatted_content)

        return formatted_content.encode("utf-8")

    def generate_table_pdf(
        self, data: list[dict], output_path: Optional[Union[str, Path]] = None
    ) -> bytes:
        """Generate PDF with table data."""
        if not REPORTLAB_AVAILABLE:
            return self._generate_fallback_table_pdf(data, output_path)

        buffer = BytesIO()

        if output_path:
            path = self._ensure_path(output_path)
            doc = SimpleDocTemplate(str(path), pagesize=self.page_size)
        else:
            doc = SimpleDocTemplate(buffer, pagesize=self.page_size)

        styles = getSampleStyleSheet()
        story = []

        # Add title
        if self.metadata.title:
            story.append(Paragraph(self.metadata.title, styles["Title"]))
            story.append(Spacer(1, 12))

        # Convert data to table
        if data:
            headers = list(data[0].keys())
            table_data = [headers]

            for row in data:
                table_data.append([str(row.get(header, "")) for header in headers])

            # Create table
            table = Table(table_data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 14),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            story.append(table)

        doc.build(story)

        if output_path:
            with open(output_path, "rb") as f:
                return f.read()
        else:
            return buffer.getvalue()

    def _generate_fallback_table_pdf(
        self, data: list[dict], output_path: Optional[Union[str, Path]]
    ) -> bytes:
        """Fallback table generation as text."""
        if not data:
            return self._generate_fallback_pdf("No data provided", output_path)

        # Format data as table
        headers = list(data[0].keys())

        # Calculate column widths
        col_widths = {header: len(header) for header in headers}
        for row in data:
            for header in headers:
                value_len = len(str(row.get(header, "")))
                if value_len > col_widths[header]:
                    col_widths[header] = value_len

        # Build table text
        content_lines = []

        # Header row
        header_row = " | ".join(header.ljust(col_widths[header]) for header in headers)
        content_lines.append(header_row)
        content_lines.append("-" * len(header_row))

        # Data rows
        for row in data:
            data_row = " | ".join(
                str(row.get(header, "")).ljust(col_widths[header]) for header in headers
            )
            content_lines.append(data_row)

        content = "\n".join(content_lines)
        return self._generate_fallback_pdf(content, output_path)


class ExcelGenerator(BaseGenerator):
    """Excel file generator with graceful degradation."""

    def __init__(self, metadata: Optional[DocumentMetadata] = None):
        super().__init__(metadata)
        self._check_dependencies()

    def _check_dependencies(self):
        """Check if required dependencies are available."""
        if not OPENPYXL_AVAILABLE:
            warnings.warn(
                "openpyxl not available - Excel generation will create CSV files instead",
                stacklevel=2,
            )

    def generate_excel(
        self,
        data: list[dict],
        output_path: Optional[Union[str, Path]] = None,
        sheet_name: str = "Data",
    ) -> bytes:
        """Generate Excel file from data."""
        if OPENPYXL_AVAILABLE:
            return self._generate_openpyxl_excel(data, output_path, sheet_name)
        else:
            return self._generate_fallback_excel(data, output_path)

    def _generate_openpyxl_excel(
        self, data: list[dict], output_path: Optional[Union[str, Path]], sheet_name: str
    ) -> bytes:
        """Generate Excel using openpyxl."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            ws["A1"] = "No data"
            if output_path:
                path = self._ensure_path(output_path)
                wb.save(path)
                with open(path, "rb") as f:
                    return f.read()
            else:
                buffer = BytesIO()
                wb.save(buffer)
                return buffer.getvalue()

        # Add headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx).value = row_data.get(header)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save or return
        if output_path:
            path = self._ensure_path(output_path)
            wb.save(path)
            with open(path, "rb") as f:
                return f.read()
        else:
            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

    def _generate_fallback_excel(
        self, data: list[dict], output_path: Optional[Union[str, Path]]
    ) -> bytes:
        """Generate CSV as fallback for Excel."""
        warnings.warn(
            "Excel generation unavailable - creating CSV file instead", stacklevel=2
        )

        csv_generator = CSVGenerator(self.metadata)
        if output_path:
            # Change extension to .csv for fallback
            path = Path(output_path).with_suffix(".csv")
            return csv_generator.generate_csv(data, path)
        else:
            return csv_generator.generate_csv(data)


class CSVGenerator(BaseGenerator):
    """CSV file generator."""

    def __init__(
        self, metadata: Optional[DocumentMetadata] = None, delimiter: str = ","
    ):
        super().__init__(metadata)
        self.delimiter = delimiter

    def generate_csv(
        self, data: list[dict], output_path: Optional[Union[str, Path]] = None
    ) -> bytes:
        """Generate CSV file from data."""
        if PANDAS_AVAILABLE and data:
            return self._generate_pandas_csv(data, output_path)
        else:
            return self._generate_manual_csv(data, output_path)

    def _generate_pandas_csv(
        self, data: list[dict], output_path: Optional[Union[str, Path]]
    ) -> bytes:
        """Generate CSV using pandas."""
        df = pd.DataFrame(data)

        if output_path:
            path = self._ensure_path(output_path)
            df.to_csv(path, index=False, sep=self.delimiter)
            with open(path, "rb") as f:
                return f.read()
        else:
            buffer = StringIO()
            df.to_csv(buffer, index=False, sep=self.delimiter)
            return buffer.getvalue().encode("utf-8")

    def _generate_manual_csv(
        self, data: list[dict], output_path: Optional[Union[str, Path]]
    ) -> bytes:
        """Generate CSV manually without pandas."""
        import csv

        if not data:
            content = "No data available\n"
            if output_path:
                path = self._ensure_path(output_path)
                with open(path, "w", encoding="utf-8") as f:
                    f.write(content)
            return content.encode("utf-8")

        # Use StringIO for in-memory CSV generation
        buffer = StringIO()

        headers = list(data[0].keys())
        writer = csv.DictWriter(buffer, fieldnames=headers, delimiter=self.delimiter)

        writer.writeheader()
        for row in data:
            writer.writerow(row)

        content = buffer.getvalue()

        if output_path:
            path = self._ensure_path(output_path)
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)

        return content.encode("utf-8")


# Convenience functions for backward compatibility
def generate_pdf(
    content: str,
    output_path: Optional[Union[str, Path]] = None,
    metadata: Optional[DocumentMetadata] = None,
) -> bytes:
    """Generate PDF document."""
    generator = PDFGenerator(metadata)
    return generator.generate_simple_pdf(content, output_path)


def generate_excel(
    data: list[dict],
    output_path: Optional[Union[str, Path]] = None,
    metadata: Optional[DocumentMetadata] = None,
) -> bytes:
    """Generate Excel document."""
    generator = ExcelGenerator(metadata)
    return generator.generate_excel(data, output_path)


def generate_csv(
    data: list[dict],
    output_path: Optional[Union[str, Path]] = None,
    metadata: Optional[DocumentMetadata] = None,
) -> bytes:
    """Generate CSV document."""
    generator = CSVGenerator(metadata)
    return generator.generate_csv(data, output_path)


# Export main classes and functions
__all__ = [
    "DocumentMetadata",
    "BaseGenerator",
    "PDFGenerator",
    "ExcelGenerator",
    "CSVGenerator",
    "generate_pdf",
    "generate_excel",
    "generate_csv",
]
