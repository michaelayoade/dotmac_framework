"""
Excel report generation for billing data.

This module provides Excel export capabilities using openpyxl.
Falls back gracefully if openpyxl is not available.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional, Union

# Optional openpyxl imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    _import_error = str(e)

    # Create stub class if openpyxl not available
    class Workbook:
        def __init__(self, *args, **kwargs):
            raise ImportError(f"openpyxl not available: {_import_error}")


class ExcelReportGenerator:
    """Excel report generator using openpyxl."""

    def __init__(self):
        """Initialize Excel generator."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                f"Excel generation requires openpyxl: {_import_error}. "
                "Install with: pip install openpyxl"
            )

    def generate_billing_report(
        self,
        report_data: dict[str, Any],
        output_path: Union[str, Path],
    ) -> str:
        """
        Generate comprehensive billing report.

        Args:
            report_data: Report data with multiple sections
            output_path: Output file path

        Returns:
            Path to generated Excel file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create sheets for different report sections
        if report_data.get('invoices'):
            self._create_invoices_sheet(wb, report_data['invoices'])

        if report_data.get('payments'):
            self._create_payments_sheet(wb, report_data['payments'])

        if report_data.get('subscriptions'):
            self._create_subscriptions_sheet(wb, report_data['subscriptions'])

        if report_data.get('usage_summary'):
            self._create_usage_sheet(wb, report_data['usage_summary'])

        # Save workbook
        wb.save(output_path)
        return str(output_path)

    def _create_invoices_sheet(self, wb: Workbook, invoices_data: list[dict]) -> None:
        """Create invoices worksheet."""
        ws = wb.create_sheet("Invoices")

        # Headers
        headers = [
            'Invoice ID', 'Invoice Number', 'Customer', 'Status',
            'Issue Date', 'Due Date', 'Subtotal', 'Tax', 'Total', 'Currency'
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Add data
        for row, invoice in enumerate(invoices_data, 2):
            ws.cell(row=row, column=1, value=str(invoice.get('id', '')))
            ws.cell(row=row, column=2, value=invoice.get('invoice_number', ''))
            ws.cell(row=row, column=3, value=invoice.get('customer_name', ''))
            ws.cell(row=row, column=4, value=invoice.get('status', ''))
            ws.cell(row=row, column=5, value=self._format_date_for_excel(invoice.get('issue_date')))
            ws.cell(row=row, column=6, value=self._format_date_for_excel(invoice.get('due_date')))
            ws.cell(row=row, column=7, value=float(invoice.get('subtotal', 0)))
            ws.cell(row=row, column=8, value=float(invoice.get('tax_amount', 0)))
            ws.cell(row=row, column=9, value=float(invoice.get('total', 0)))
            ws.cell(row=row, column=10, value=invoice.get('currency', 'USD'))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _create_payments_sheet(self, wb: Workbook, payments_data: list[dict]) -> None:
        """Create payments worksheet."""
        ws = wb.create_sheet("Payments")

        # Headers
        headers = [
            'Payment ID', 'Invoice ID', 'Customer', 'Amount',
            'Currency', 'Date', 'Method', 'Status', 'Transaction ID'
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Add data
        for row, payment in enumerate(payments_data, 2):
            ws.cell(row=row, column=1, value=str(payment.get('id', '')))
            ws.cell(row=row, column=2, value=str(payment.get('invoice_id', '')))
            ws.cell(row=row, column=3, value=payment.get('customer_name', ''))
            ws.cell(row=row, column=4, value=float(payment.get('amount', 0)))
            ws.cell(row=row, column=5, value=payment.get('currency', 'USD'))
            ws.cell(row=row, column=6, value=self._format_date_for_excel(payment.get('payment_date')))
            ws.cell(row=row, column=7, value=payment.get('payment_method', ''))
            ws.cell(row=row, column=8, value=payment.get('status', ''))
            ws.cell(row=row, column=9, value=payment.get('transaction_id', ''))

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_subscriptions_sheet(self, wb: Workbook, subscriptions_data: list[dict]) -> None:
        """Create subscriptions worksheet."""
        ws = wb.create_sheet("Subscriptions")

        # Headers
        headers = [
            'Subscription ID', 'Customer', 'Plan', 'Status', 'Start Date',
            'Current Period Start', 'Current Period End', 'Next Billing',
            'Monthly Revenue', 'Billing Cycle'
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Add data
        for row, subscription in enumerate(subscriptions_data, 2):
            ws.cell(row=row, column=1, value=str(subscription.get('id', '')))
            ws.cell(row=row, column=2, value=subscription.get('customer_name', ''))
            ws.cell(row=row, column=3, value=subscription.get('plan_name', ''))
            ws.cell(row=row, column=4, value=subscription.get('status', ''))
            ws.cell(row=row, column=5, value=self._format_date_for_excel(subscription.get('start_date')))
            ws.cell(row=row, column=6, value=self._format_date_for_excel(subscription.get('current_period_start')))
            ws.cell(row=row, column=7, value=self._format_date_for_excel(subscription.get('current_period_end')))
            ws.cell(row=row, column=8, value=self._format_date_for_excel(subscription.get('next_billing_date')))
            ws.cell(row=row, column=9, value=float(subscription.get('monthly_price', 0)))
            ws.cell(row=row, column=10, value=subscription.get('billing_cycle', ''))

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_usage_sheet(self, wb: Workbook, usage_data: list[dict]) -> None:
        """Create usage summary worksheet."""
        ws = wb.create_sheet("Usage Summary")

        # Headers
        headers = [
            'Subscription ID', 'Customer', 'Usage Date', 'Meter Type',
            'Quantity', 'Unit', 'Billable Amount'
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Add data
        for row, usage in enumerate(usage_data, 2):
            ws.cell(row=row, column=1, value=str(usage.get('subscription_id', '')))
            ws.cell(row=row, column=2, value=usage.get('customer_name', ''))
            ws.cell(row=row, column=3, value=self._format_date_for_excel(usage.get('usage_date')))
            ws.cell(row=row, column=4, value=usage.get('meter_type', ''))
            ws.cell(row=row, column=5, value=float(usage.get('quantity', 0)))
            ws.cell(row=row, column=6, value=usage.get('unit', ''))
            ws.cell(row=row, column=7, value=float(usage.get('billable_amount', 0)))

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _auto_adjust_columns(self, worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _format_date_for_excel(self, date_value: Any) -> Optional[datetime]:
        """Format date value for Excel."""
        if date_value is None:
            return None

        if isinstance(date_value, datetime):
            return date_value

        if isinstance(date_value, date):
            return datetime.combine(date_value, datetime.min.time())

        # Try to parse string dates
        if isinstance(date_value, str):
            try:
                return datetime.fromisoformat(date_value.replace('Z', '+00:00'))
            except:
                return None

        return None
